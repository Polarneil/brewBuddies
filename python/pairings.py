import openpyxl
import random
import os


def read_simple_xlsx(filepath):
    """Reads an XLSX file and returns the rows as a list of tuples, skipping the first row (headers),
    only including rows where the value in column G is "TRUE"."""
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        # Read relevant columns: Name (A), Email (B), Geo Team (E), and Pairing Status (G)
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 7 and str(row[6]).upper() == "TRUE":  # Check if column G is "TRUE" (case-insensitive)
                rows.append((row[0], row[1], row[4]))  # Name, Email, Geo Team
        return rows
    except FileNotFoundError:
        print(f"File not found: {filepath}")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []


def create_random_pairings(employees_data, existing_pairs):
    """Creates random pairings from the list of employee data (name, email, Geo Team),
    avoiding existing pairs and ensuring people in a pair have different Geo Teams.
    If the number of employees is odd, it will attempt to create one group of three."""
    random.shuffle(employees_data)
    pairings = []
    used_indices = set()
    n = len(employees_data)

    # Attempt to form pairs first
    for i in range(n):
        if i not in used_indices:
            potential_partner_index = -1
            for j in range(i + 1, n):
                if j not in used_indices and employees_data[i][2] != employees_data[j][2]:
                    pair = tuple(sorted((employees_data[i][0], employees_data[j][0])))
                    if pair not in existing_pairs:
                        potential_partner_index = j
                        break

            if potential_partner_index != -1:
                pairings.append(((employees_data[i][0], employees_data[i][1]),
                                 (employees_data[potential_partner_index][0], employees_data[potential_partner_index][1]),
                                 (employees_data[i][2], employees_data[potential_partner_index][2])))
                used_indices.add(i)
                used_indices.add(potential_partner_index)

    # Handle remaining individuals to form a group of three if the initial count was odd
    remaining_indices = [i for i in range(n) if i not in used_indices]
    if len(remaining_indices) == 3:
        group_names = tuple(sorted((employees_data[remaining_indices[0]][0],
                                    employees_data[remaining_indices[1]][0],
                                    employees_data[remaining_indices[2]][0])))
        # Basic check for existing triplets (can be improved if needed)
        is_new_triplet = True
        for pair in [(group_names[0], group_names[1]), (group_names[0], group_names[2]), (group_names[1], group_names[2])]:
            if pair in existing_pairs:
                is_new_triplet = False
                break
        if is_new_triplet:
            pairings.append(((employees_data[remaining_indices[0]][0], employees_data[remaining_indices[0]][1]),
                             (employees_data[remaining_indices[1]][0], employees_data[remaining_indices[1]][1]),
                             (employees_data[remaining_indices[2]][0], employees_data[remaining_indices[2]][1]),
                             (employees_data[remaining_indices[0]][2], employees_data[remaining_indices[1]][2], employees_data[remaining_indices[2]][2])))

    elif len(remaining_indices) == 2:
        pair_names = tuple(sorted((employees_data[remaining_indices[0]][0], employees_data[remaining_indices[1]][0])))
        if pair_names not in existing_pairs:
            pairings.append(((employees_data[remaining_indices[0]][0], employees_data[remaining_indices[0]][1]),
                             (employees_data[remaining_indices[1]][0], employees_data[remaining_indices[1]][1]),
                             (employees_data[remaining_indices[0]][2], employees_data[remaining_indices[1]][2])))
        else:
            # If the remaining pair exists, they are left as individuals (can be adjusted)
            pairings.append(((employees_data[remaining_indices[0]][0], employees_data[remaining_indices[0]][1]),
                             (employees_data[remaining_indices[0]][2],)))
            pairings.append(((employees_data[remaining_indices[1]][0], employees_data[remaining_indices[1]][1]),
                             (employees_data[remaining_indices[1]][2],)))
    elif len(remaining_indices) == 1:
        pairings.append(((employees_data[remaining_indices[0]][0], employees_data[remaining_indices[0]][1]),
                         (employees_data[remaining_indices[0]][2],)))

    return pairings


def save_pairings_to_xlsx(pairings, filepath):
    """Saves the pairings to a new XLSX file, including email and Geo Team."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, group in enumerate(pairings, start=1):
        col = 1
        num_people = len(group) - (1 if len(group[-1]) > 1 or isinstance(group[-1], str) else 0)
        geo_teams_start_col = num_people * 2 + 1

        for j in range(num_people):
            person_info = group[j]
            sheet.cell(row=i, column=col, value=person_info[0])  # Name
            sheet.cell(row=i, column=col + 1, value=person_info[1])  # Email
            col += 2

        # Add Geo Teams
        geo_teams = group[-1]
        if isinstance(geo_teams, tuple):
            for k, geo_team in enumerate(geo_teams):
                sheet.cell(row=i, column=geo_teams_start_col + k, value=geo_team)
        elif isinstance(geo_teams, str):
            sheet.cell(row=i, column=geo_teams_start_col, value=geo_teams)

    workbook.save(filepath)


def get_next_available_filename(directory, base_name, extension):
    """Finds the next available filename in the specified directory."""
    i = 1
    while True:
        filename = f"{base_name}_{i}.{extension}"
        if not os.path.exists(os.path.join(directory, filename)):
            return os.path.join(directory, filename)
        i += 1


def read_existing_pairs(directory, base_name, extension):
    """Reads all existing pairing files and returns a set of existing pairs (based on names)."""
    existing_pairs = set()
    i = 1
    while True:
        filename = f"{base_name}_{i}.{extension}"
        filepath = os.path.join(directory, filename)
        if not os.path.exists(filepath):
            break
        try:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                names = tuple(sorted(filter(None, row[:sheet.max_column // 2 * 2:2]))) # Extract names from even columns
                if len(names) == 2:
                    existing_pairs.add(names)
                elif len(names) == 3:
                    existing_pairs.add(tuple(sorted((names[0], names[1]))))
                    existing_pairs.add(tuple(sorted((names[0], names[2]))))
                    existing_pairs.add(tuple(sorted((names[1], names[2]))))
        except Exception as e:
            print(f"Error reading existing pairings file {filename}: {e}")
        i += 1
    return existing_pairs


# Example usage:
xlsx_path = "../data/data_members.xlsx"
employees_data = read_simple_xlsx(xlsx_path)
if employees_data:
    output_dir = "../data/pairings"
    os.makedirs(output_dir, exist_ok=True)
    all_existing_pairs = read_existing_pairs(output_dir, "pairing", "xlsx")
    all_pairings = create_random_pairings(employees_data, all_existing_pairs)
    output_path = get_next_available_filename(output_dir, "pairing", "xlsx")
    save_pairings_to_xlsx(all_pairings, output_path)
    print(f"Pairings saved to: {output_path}")
